"""
Aplicación web completa para el sistema de gestión de laboratorios.
Utiliza Streamlit para crear una interfaz web con todos los CRUDs.
"""

import streamlit as st
import pandas as pd
from datetime import datetime
import os

from models.database import inicializar_bd
from managers.materia_manager import MateriaManager
from managers.paralelo_manager import ParaleloManager
from managers.estudiante_manager import EstudianteManager
from managers.laboratorio_manager import LaboratorioManager
from managers.calificacion_manager import CalificacionManager
from utils.pdf_exporter import PDFExporter

# Configuración de la página
st.set_page_config(
    page_title="Sistema de Gestión de Laboratorios",
    page_icon="\U0001F393",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personalizado
st.markdown("""
<style>
.main-header {
    font-size: 2.5rem;
    font-weight: bold;
    color: #1f77b4;
    text-align: center;
    margin-bottom: 1rem;
    padding: 1rem;
    background: linear-gradient(90deg, #e3f2fd 0%, #bbdefb 100%);
    border-radius: 10px;
    border-left: 5px solid #1f77b4;
}

.sub-header {
    font-size: 1.2rem;
    color: #ff7f0e;
    text-align: center;
    margin-bottom: 2rem;
    font-style: italic;
}

.metric-card {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    padding: 1rem;
    border-radius: 10px;
    color: white;
    text-align: center;
    margin: 0.5rem 0;
}

.stTabs [data-baseweb="tab-list"] {
    gap: 2px;
}

.stTabs [data-baseweb="tab"] {
    height: 50px;
    padding-left: 20px;
    padding-right: 20px;
}
</style>
""", unsafe_allow_html=True)

def inicializar_aplicacion():
    """Inicializa la aplicación y la base de datos"""
    if 'initialized' not in st.session_state:
        inicializar_bd()
        st.session_state.initialized = True

def mostrar_titulo_principal():
    """Muestra el título principal del sistema"""
    st.markdown("""
    <div class="main-header">
        Sistema de Gestión de Laboratorios
    </div>
    <div class="sub-header">
        Away - Gestión de Laboratorios
    </div>
    """, unsafe_allow_html=True)

def sidebar_navegacion():
    """Crea la barra lateral de navegación"""
    with st.sidebar:
        st.title("Navegación")
        
        pagina = st.selectbox(
            "Seleccionar módulo:",
            ["Dashboard", "Materias", "Paralelos", "Estudiantes", 
             "Laboratorios", "Calificaciones", "Reportes", "Estadísticas"]
        )
        
        st.markdown("---")
        st.markdown("### Información del Sistema")
        st.markdown("""
        **Away - Sistema de Gestión de Laboratorios**
        - Desarrollado por: ErwinSaul
        """)
        
        st.markdown("---")
        st.markdown("### Sesión Actual")
        st.write(f"Hora: {datetime.now().strftime('%H:%M:%S')}")
        st.write(f"Fecha: {datetime.now().strftime('%d/%m/%Y')}")
        
    return pagina

def mostrar_dashboard():
    """Muestra el dashboard principal"""
    st.header("Dashboard del Sistema")
    
    # Obtener estadísticas
    try:
        stats = MateriaManager.obtener_estadisticas_generales()
        
        # Mostrar métricas principales
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric(
                label="Total Materias",
                value=stats.get('total_materias', 0)
            )
        
        with col2:
            st.metric(
                label="Total Paralelos", 
                value=stats.get('total_paralelos', 0)
            )
        
        with col3:
            st.metric(
                label="Total Estudiantes",
                value=stats.get('total_estudiantes', 0)
            )
        
        with col4:
            st.metric(
                label="Total Laboratorios",
                value=stats.get('total_laboratorios', 0)
            )
        
        st.markdown("---")
        
        # Sección de acciones rápidas
        st.subheader("Acciones Rápidas")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            if st.button("Nueva Materia", use_container_width=True, type="primary"):
                st.session_state.pagina = "Materias"
                st.rerun()
        
        with col2:
            if st.button("Nuevo Estudiante", use_container_width=True):
                st.session_state.pagina = "Estudiantes"
                st.rerun()
        
        with col3:
            if st.button("Nueva Calificación", use_container_width=True):
                st.session_state.pagina = "Calificaciones"
                st.rerun()
        
        with col4:
            if st.button("Ver Reportes", use_container_width=True):
                st.session_state.pagina = "Reportes"
                st.rerun()
        
        st.markdown("---")
        
        # Resumen de materias
        st.subheader("Materias Registradas")
        
        materias = MateriaManager.listar_materias()
        
        if materias:
            datos = []
            for materia in materias:
                datos.append({
                    'ID': materia.id,
                    'Sigla': materia.sigla,
                    'Materia': materia.materia,
                    'Paralelos': materia.contar_paralelos(),
                    'Estudiantes': materia.contar_estudiantes_total(),
                    'Laboratorios': materia.contar_laboratorios()
                })
            
            df = pd.DataFrame(datos)
            st.dataframe(
                df,
                use_container_width=True,
                hide_index=True
            )
        else:
            st.info("No hay materias registradas. Crea tu primera materia para comenzar!")
        
    except Exception as e:
        st.error(f"Error al cargar dashboard: {e}")

def pagina_materias():
    """Página de gestión de materias"""
    st.header("Gestión de Materias")
    
    tab1, tab2, tab3, tab4 = st.tabs(["Lista de Materias", "Nueva Materia", "Editar Materia", "Estadísticas"])
    
    with tab1:
        st.subheader("Materias Registradas")
        
        col1, col2 = st.columns([3, 1])
        with col2:
            if st.button("Actualizar", type="secondary"):
                st.rerun()
        
        try:
            materias = MateriaManager.listar_materias()
            
            if materias:
                # Buscador
                busqueda = st.text_input("Buscar materias:", placeholder="Ingrese término de búsqueda")
                
                if busqueda:
                    materias_filtradas = MateriaManager.buscar_materias(busqueda)
                else:
                    materias_filtradas = materias
                
                datos = []
                for materia in materias_filtradas:
                    datos.append({
                        'ID': materia.id,
                        'Sigla': materia.sigla,
                        'Materia': materia.materia,
                        'Paralelos': materia.contar_paralelos(),
                        'Estudiantes': materia.contar_estudiantes_total(),
                        'Laboratorios': materia.contar_laboratorios()
                    })
                
                df = pd.DataFrame(datos)
                
                # Selector para edición/eliminación
                if len(datos) > 0:
                    st.dataframe(df, use_container_width=True, hide_index=True)
                    
                    st.subheader("Acciones sobre Materias")
                    
                    # Selector de materia
                    opciones_materias = {f"{m.sigla} - {m.materia}": m.id for m in materias_filtradas}
                    if opciones_materias:
                        materia_seleccionada = st.selectbox(
                            "Seleccionar materia para acciones:",
                            options=list(opciones_materias.keys())
                        )
                        
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            # Manejo de estado de confirmación
                            if 'confirmar_eliminar_materia' not in st.session_state:
                                st.session_state.confirmar_eliminar_materia = None

                            materia_id = opciones_materias[materia_seleccionada]

                            if st.session_state.confirmar_eliminar_materia == materia_id:
                                st.warning("⚠️ ¿Está seguro? Esta acción eliminará todos los datos relacionados.")
                                col_si, col_no = st.columns(2)

                                with col_si:
                                    if st.button("✓ Sí, eliminar", type="primary", use_container_width=True, key=f"confirmar_si_{materia_id}"):
                                        resultado = MateriaManager.eliminar_materia(materia_id, forzar=True)

                                        if resultado['success']:
                                            st.success(resultado['mensaje'])
                                            st.session_state.confirmar_eliminar_materia = None
                                            st.rerun()
                                        else:
                                            st.error(resultado['mensaje'])
                                            st.session_state.confirmar_eliminar_materia = None

                                with col_no:
                                    if st.button("✗ Cancelar", use_container_width=True, key=f"confirmar_no_{materia_id}"):
                                        st.session_state.confirmar_eliminar_materia = None
                                        st.rerun()
                            else:
                                if st.button("Eliminar Materia", type="primary", use_container_width=True):
                                    st.session_state.confirmar_eliminar_materia = materia_id
                                    st.rerun()
                        
                        with col2:
                            if st.button("Ver Detalles", use_container_width=True):
                                materia_id = opciones_materias[materia_seleccionada]
                                materia = MateriaManager.obtener_materia(materia_id)
                                
                                if materia:
                                    stats = materia.estadisticas_completas()
                                    
                                    st.info(f"""
                                    **Materia:** {stats['materia']}
                                    **Sigla:** {stats['sigla']}
                                    **Paralelos:** {stats['paralelos']}
                                    **Estudiantes:** {stats['estudiantes']}
                                    **Laboratorios:** {stats['laboratorios']}
                                    """)
                else:
                    st.info("No se encontraron materias")
            else:
                st.info("No hay materias registradas")
                
        except Exception as e:
            st.error(f"Error al cargar materias: {e}")
    
    with tab2:
        st.subheader("Crear Nueva Materia")
        
        with st.form("form_nueva_materia"):
            col1, col2 = st.columns(2)

            with col1:
                materia = st.text_input("Nombre de la Materia *",
                                       placeholder="Ej: Programación I")

            with col2:
                sigla = st.text_input("Sigla *",
                                     placeholder="Ej: SIS-111")

            submitted = st.form_submit_button("Crear Materia",
                                            type="primary",
                                            use_container_width=True)

            if submitted:
                if not materia or not sigla:
                    st.error("Todos los campos son obligatorios")
                else:
                    resultado = MateriaManager.crear_materia(materia, sigla)
                    
                    if resultado:
                        st.success(f"Materia {sigla} creada exitosamente")
                        st.rerun()
                    else:
                        st.error("No se pudo crear la materia (sigla duplicada?)")
    
    with tab3:
        st.subheader("Editar Materia")
        
        materias = MateriaManager.listar_materias()
        
        if materias:
            opciones_materias = {f"{m.sigla} - {m.materia}": m for m in materias}
            materia_seleccionada = st.selectbox(
                "Seleccionar materia a editar:",
                options=list(opciones_materias.keys())
            )
            
            if materia_seleccionada:
                materia = opciones_materias[materia_seleccionada]

                with st.form("form_editar_materia"):
                    col1, col2 = st.columns(2)

                    with col1:
                        nuevo_nombre = st.text_input("Nombre de la Materia", value=materia.materia)

                    with col2:
                        nueva_sigla = st.text_input("Sigla", value=materia.sigla)

                    submitted = st.form_submit_button("Actualizar Materia",
                                                    type="primary",
                                                    use_container_width=True)

                    if submitted:
                        if not nuevo_nombre or not nueva_sigla:
                            st.error("Todos los campos son obligatorios")
                        else:
                            resultado = MateriaManager.actualizar_materia(
                                materia.id,
                                materia=nuevo_nombre,
                                sigla=nueva_sigla
                            )
                            
                            if resultado:
                                st.success("Materia actualizada exitosamente")
                                st.rerun()
                            else:
                                st.error("No se pudo actualizar la materia")
        else:
            st.info("No hay materias registradas para editar")
    
    with tab4:
        st.subheader("Estadísticas de Materias")
        
        try:
            stats = MateriaManager.obtener_estadisticas_generales()
            materias = MateriaManager.listar_materias()
            
            # Métricas generales
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("Total Materias", stats['total_materias'])
            
            with col2:
                st.metric("Total Paralelos", stats['total_paralelos'])
            
            with col3:
                st.metric("Total Estudiantes", stats['total_estudiantes'])
            
            # Detalles por materia
            st.subheader("Detalle por Materia")
            
            if materias:
                datos_stats = []
                for materia in materias:
                    stats_materia = materia.estadisticas_completas()
                    datos_stats.append({
                        'Sigla': stats_materia['sigla'],
                        'Materia': stats_materia['materia'],
                        'Paralelos': stats_materia['paralelos'],
                        'Estudiantes': stats_materia['estudiantes'],
                        'Laboratorios': stats_materia['laboratorios']
                    })
                
                df_stats = pd.DataFrame(datos_stats)
                st.dataframe(df_stats, use_container_width=True, hide_index=True)
            
        except Exception as e:
            st.error(f"Error al cargar estadísticas: {e}")

def pagina_paralelos():
    """Página de gestión de paralelos"""
    st.header("Gestión de Paralelos")
    
    tab1, tab2, tab3, tab4 = st.tabs(["Lista de Paralelos", "Nuevo Paralelo", "Editar Paralelo", "Estadísticas"])
    
    with tab1:
        st.subheader("Paralelos por Materia")
        
        # Selector de materia
        materias = MateriaManager.listar_materias()
        
        if not materias:
            st.warning("No hay materias registradas. Debe crear materias primero.")
            return
        
        opciones_materias = {f"{m.sigla} - {m.materia}": m.id for m in materias}
        materia_seleccionada = st.selectbox(
            "Seleccionar materia:",
            options=list(opciones_materias.keys())
        )
        
        if materia_seleccionada:
            materia_id = opciones_materias[materia_seleccionada]
            paralelos = ParaleloManager.listar_paralelos_por_materia(materia_id)
            
            if paralelos:
                datos = []
                for paralelo in paralelos:
                    datos.append({
                        'ID': paralelo.id,
                        'Paralelo': paralelo.paralelo,
                        'Docente de Teoría': paralelo.docente_teoria,
                        'Estudiantes': paralelo.contar_estudiantes(),
                        'Grupos': paralelo.contar_grupos(),
                        'Promedio': f"{paralelo.promedio_general():.2f}"
                    })

                df = pd.DataFrame(datos)
                st.dataframe(df, use_container_width=True, hide_index=True)
                
                # Acciones sobre paralelos
                st.subheader("Acciones")
                
                opciones_paralelos = {f"Paralelo {p.paralelo}": p.id for p in paralelos}
                paralelo_seleccionado = st.selectbox(
                    "Seleccionar paralelo:",
                    options=list(opciones_paralelos.keys())
                )
                
                col1, col2 = st.columns(2)
                
                with col1:
                    # Manejo de estado de confirmación
                    if 'confirmar_eliminar_paralelo' not in st.session_state:
                        st.session_state.confirmar_eliminar_paralelo = None

                    paralelo_id = opciones_paralelos[paralelo_seleccionado]

                    if st.session_state.confirmar_eliminar_paralelo == paralelo_id:
                        st.warning("⚠️ ¿Está seguro de eliminar este paralelo?")
                        col_si, col_no = st.columns(2)

                        with col_si:
                            if st.button("✓ Sí, eliminar", type="primary", use_container_width=True, key=f"conf_par_si_{paralelo_id}"):
                                resultado = ParaleloManager.eliminar_paralelo(paralelo_id, forzar=True)

                                if resultado['success']:
                                    st.success(resultado['mensaje'])
                                    st.session_state.confirmar_eliminar_paralelo = None
                                    st.rerun()
                                else:
                                    st.error(resultado['mensaje'])
                                    st.session_state.confirmar_eliminar_paralelo = None

                        with col_no:
                            if st.button("✗ Cancelar", use_container_width=True, key=f"conf_par_no_{paralelo_id}"):
                                st.session_state.confirmar_eliminar_paralelo = None
                                st.rerun()
                    else:
                        if st.button("Eliminar Paralelo", type="primary", use_container_width=True):
                            st.session_state.confirmar_eliminar_paralelo = paralelo_id
                            st.rerun()
                
                with col2:
                    if st.button("Ver Estadísticas"):
                        paralelo_id = opciones_paralelos[paralelo_seleccionado]
                        stats = EstudianteManager.obtener_estadisticas_paralelo(paralelo_id)
                        
                        if 'error' not in stats:
                            st.info(f"""
                            **Paralelo:** {stats['paralelo_info']}
                            **Estudiantes:** {stats['total_estudiantes']}
                            **Grupos:** {stats['total_grupos']}
                            **Sin grupo:** {stats['estudiantes_sin_grupo']}
                            **Promedio:** {stats['promedio_general']:.2f}
                            """)
                        else:
                            st.error(stats['error'])
            else:
                st.info("No hay paralelos registrados para esta materia")
    
    with tab2:
        st.subheader("Crear Nuevo Paralelo")
        
        materias = MateriaManager.listar_materias()
        
        if materias:
            with st.form("form_nuevo_paralelo"):
                opciones_materias = {f"{m.sigla} - {m.materia}": m.id for m in materias}
                materia_seleccionada = st.selectbox(
                    "Materia:",
                    options=list(opciones_materias.keys())
                )

                col1, col2 = st.columns(2)

                with col1:
                    paralelo = st.text_input("Nombre del paralelo *", placeholder="A, B, C, etc.")

                with col2:
                    docente_teoria = st.text_input("Docente de Teoría *", placeholder="Ej: Ing. Juan Pérez")

                submitted = st.form_submit_button("Crear Paralelo",
                                                type="primary",
                                                use_container_width=True)

                if submitted:
                    if not paralelo or not docente_teoria:
                        st.error("Todos los campos son obligatorios")
                    else:
                        materia_id = opciones_materias[materia_seleccionada]
                        resultado = ParaleloManager.crear_paralelo(materia_id, paralelo, docente_teoria)

                        if resultado:
                            st.success(f"Paralelo {paralelo} creado exitosamente")
                            st.rerun()
                        else:
                            st.error("No se pudo crear el paralelo")
        else:
            st.warning("No hay materias registradas. Debe crear materias primero.")

    with tab3:
        st.subheader("Editar Paralelo")

        # Obtener todos los paralelos agrupados por materia
        materias = MateriaManager.listar_materias()

        if not materias:
            st.warning("No hay materias registradas.")
            return

        # Crear lista de paralelos disponibles
        paralelos_disponibles = []
        for materia in materias:
            paralelos = ParaleloManager.listar_paralelos_por_materia(materia.id)
            for paralelo in paralelos:
                paralelos_disponibles.append({
                    'texto': f"{materia.sigla} - Paralelo {paralelo.paralelo}",
                    'paralelo': paralelo
                })

        if paralelos_disponibles:
            opciones_paralelos = {p['texto']: p['paralelo'] for p in paralelos_disponibles}
            paralelo_seleccionado = st.selectbox(
                "Seleccionar paralelo a editar:",
                options=list(opciones_paralelos.keys())
            )

            if paralelo_seleccionado:
                paralelo = opciones_paralelos[paralelo_seleccionado]

                with st.form("form_editar_paralelo"):
                    col1, col2 = st.columns(2)

                    with col1:
                        nuevo_nombre = st.text_input("Nombre del paralelo", value=paralelo.paralelo)

                    with col2:
                        nuevo_docente = st.text_input("Docente de Teoría", value=paralelo.docente_teoria)

                    submitted = st.form_submit_button("Actualizar Paralelo",
                                                    type="primary",
                                                    use_container_width=True)

                    if submitted:
                        if not nuevo_nombre or not nuevo_docente:
                            st.error("Todos los campos son obligatorios")
                        else:
                            resultado = ParaleloManager.actualizar_paralelo(
                                paralelo.id,
                                paralelo=nuevo_nombre,
                                docente_teoria=nuevo_docente
                            )

                            if resultado:
                                st.success("Paralelo actualizado exitosamente")
                                st.rerun()
                            else:
                                st.error("No se pudo actualizar el paralelo")
        else:
            st.info("No hay paralelos registrados para editar")

    with tab4:
        st.subheader("Estadísticas de Paralelos")
        
        materias = MateriaManager.listar_materias()
        
        for materia in materias:
            paralelos = ParaleloManager.listar_paralelos_por_materia(materia.id)
            
            if paralelos:
                st.write(f"**{materia.sigla} - {materia.materia}**")
                
                datos_stats = []
                for paralelo in paralelos:
                    datos_stats.append({
                        'Paralelo': paralelo.paralelo,
                        'Estudiantes': paralelo.contar_estudiantes(),
                        'Grupos': paralelo.contar_grupos(),
                        'Promedio': f"{paralelo.promedio_general():.2f}"
                    })
                
                df_stats = pd.DataFrame(datos_stats)
                st.dataframe(df_stats, use_container_width=True, hide_index=True)
                
                st.markdown("---")

def pagina_estudiantes():
    """Página de gestión de estudiantes"""
    st.header("Gestión de Estudiantes")
    
    tab1, tab2, tab3, tab4 = st.tabs(["Lista de Estudiantes", "Nuevo Estudiante", "Organizar Grupos", "Búsqueda"])
    
    with tab1:
        st.subheader("Estudiantes por Paralelo")
        
        # Obtener paralelos
        paralelos_disponibles = []
        materias = MateriaManager.listar_materias()
        
        for materia in materias:
            paralelos = ParaleloManager.listar_paralelos_por_materia(materia.id)
            for paralelo in paralelos:
                paralelos_disponibles.append({
                    'texto': f"{materia.sigla} - Paralelo {paralelo.paralelo}",
                    'paralelo_id': paralelo.id
                })
        
        if not paralelos_disponibles:
            st.warning("No hay paralelos registrados. Debe crear paralelos primero.")
            return
        
        opciones_paralelos = {p['texto']: p['paralelo_id'] for p in paralelos_disponibles}
        paralelo_seleccionado = st.selectbox(
            "Seleccionar paralelo:",
            options=list(opciones_paralelos.keys())
        )
        
        if paralelo_seleccionado:
            paralelo_id = opciones_paralelos[paralelo_seleccionado]
            estudiantes = EstudianteManager.listar_por_paralelo(paralelo_id)
            
            if estudiantes:
                datos = []
                for estudiante in estudiantes:
                    datos.append({
                        'ID': estudiante.id,
                        'CI': estudiante.ci,
                        'Nombre': estudiante.nombre,
                        'Grupo': estudiante.grupo or "Sin asignar",
                        'Calificaciones': estudiante.contar_calificaciones(),
                        'Promedio': f"{estudiante.promedio_calificaciones():.2f}"
                    })
                
                df = pd.DataFrame(datos)
                st.dataframe(df, use_container_width=True, hide_index=True)
                
                # Acciones sobre estudiantes
                st.subheader("Acciones sobre Estudiantes")
                
                opciones_estudiantes = {f"{e.ci} - {e.nombre}": e.id for e in estudiantes}
                estudiante_seleccionado = st.selectbox(
                    "Seleccionar estudiante:",
                    options=list(opciones_estudiantes.keys())
                )
                
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    if st.button("Ver Detalles", use_container_width=True):
                        estudiante_id = opciones_estudiantes[estudiante_seleccionado]
                        estudiante = EstudianteManager.obtener_estudiante(estudiante_id)
                        
                        if estudiante:
                            calificaciones = estudiante.calificaciones_por_laboratorio()
                            
                            info = f"""
                            **Nombre:** {estudiante.nombre}
                            **CI:** {estudiante.ci}
                            **Grupo:** {estudiante.grupo or 'Sin asignar'}
                            **Promedio:** {estudiante.promedio_calificaciones():.2f}
                            **Calificaciones:** {estudiante.contar_calificaciones()}
                            """
                            
                            st.info(info)
                            
                            if calificaciones:
                                st.write("**Calificaciones por laboratorio:**")
                                for lab, datos in calificaciones.items():
                                    st.write(f"- {lab}: {datos['calificacion']:.1f} - {datos['titulo']}")
                
                with col2:
                    if st.button("Editar Estudiante", use_container_width=True):
                        estudiante_id = opciones_estudiantes[estudiante_seleccionado]
                        st.session_state['editar_estudiante_id'] = estudiante_id
                        
                with col3:
                    # Manejo de estado de confirmación
                    if 'confirmar_eliminar_estudiante' not in st.session_state:
                        st.session_state.confirmar_eliminar_estudiante = None

                    estudiante_id = opciones_estudiantes[estudiante_seleccionado]

                    if st.session_state.confirmar_eliminar_estudiante == estudiante_id:
                        st.warning("⚠️ ¿Está seguro de eliminar este estudiante?")
                        col_si, col_no = st.columns(2)

                        with col_si:
                            if st.button("✓ Sí, eliminar", type="primary", use_container_width=True, key=f"conf_est_si_{estudiante_id}"):
                                resultado = EstudianteManager.eliminar_estudiante(estudiante_id, forzar=True)

                                if resultado['success']:
                                    st.success(resultado['mensaje'])
                                    st.session_state.confirmar_eliminar_estudiante = None
                                    st.rerun()
                                else:
                                    st.error(resultado['mensaje'])
                                    st.session_state.confirmar_eliminar_estudiante = None

                        with col_no:
                            if st.button("✗ Cancelar", use_container_width=True, key=f"conf_est_no_{estudiante_id}"):
                                st.session_state.confirmar_eliminar_estudiante = None
                                st.rerun()
                    else:
                        if st.button("Eliminar Estudiante", type="primary", use_container_width=True):
                            st.session_state.confirmar_eliminar_estudiante = estudiante_id
                            st.rerun()
            else:
                st.info("No hay estudiantes registrados en este paralelo")
    
    with tab2:
        st.subheader("Registrar Nuevo Estudiante")

        # Obtener paralelos
        paralelos_disponibles = []
        materias = MateriaManager.listar_materias()

        for materia in materias:
            paralelos = ParaleloManager.listar_paralelos_por_materia(materia.id)
            for paralelo in paralelos:
                paralelos_disponibles.append({
                    'texto': f"{materia.sigla} - Paralelo {paralelo.paralelo}",
                    'paralelo_id': paralelo.id
                })

        if paralelos_disponibles:
            with st.form("form_nuevo_estudiante"):
                opciones_paralelos = {p['texto']: p['paralelo_id'] for p in paralelos_disponibles}
                paralelo_seleccionado = st.selectbox(
                    "Paralelo:",
                    options=list(opciones_paralelos.keys())
                )

                col1, col2 = st.columns(2)

                with col1:
                    nombre = st.text_input("Nombre completo *", placeholder="Juan Pérez")
                    ci = st.text_input("Cédula de identidad *", placeholder="12345678")

                with col2:
                    grupo = st.text_input("Grupo (opcional)", placeholder="Grupo 1")

                submitted = st.form_submit_button("Registrar Estudiante",
                                                type="primary",
                                                use_container_width=True)

                if submitted:
                    if not nombre or not ci:
                        st.error("Nombre y CI son obligatorios")
                    else:
                        paralelo_id = opciones_paralelos[paralelo_seleccionado]
                        ci_limpio = ci.strip().upper()

                        # Verificar si el CI ya existe en otros paralelos
                        estudiantes_existentes = EstudianteManager.buscar_todos_por_ci(ci_limpio)

                        # Verificar si ya está inscrito en este paralelo específico
                        estudiante_en_paralelo = EstudianteManager.buscar_por_ci_en_paralelo(ci_limpio, paralelo_id)

                        if estudiante_en_paralelo:
                            st.error(f"El estudiante con CI {ci} ya está inscrito en este paralelo")
                        else:
                            # Mostrar advertencia si ya existe en otros paralelos
                            if estudiantes_existentes:
                                st.warning(f"ℹ️ El CI {ci_limpio} ya está registrado en {len(estudiantes_existentes)} materia(s). Se inscribirá en una nueva materia/paralelo.")

                            resultado = EstudianteManager.registrar_estudiante(
                                nombre, ci, paralelo_id, grupo or None
                            )

                            if resultado is None:
                                st.error("Error interno: El sistema no pudo procesar la inscripción")
                            elif resultado['success']:
                                if estudiantes_existentes:
                                    st.success(f"Estudiante {nombre} inscrito exitosamente en una nueva materia")
                                else:
                                    st.success(resultado['mensaje'])
                                st.rerun()
                            else:
                                st.error(f"No se pudo registrar el estudiante: {resultado['mensaje']}")
        else:
            st.warning("No hay paralelos registrados. Debe crear paralelos primero.")
        
        # Formulario de edición si hay un estudiante seleccionado
        if 'editar_estudiante_id' in st.session_state:
            st.subheader("Editar Estudiante")
            
            estudiante_id = st.session_state['editar_estudiante_id']
            estudiante = EstudianteManager.obtener_estudiante(estudiante_id)
            
            if estudiante:
                with st.form("form_editar_estudiante"):
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        nuevo_nombre = st.text_input("Nombre completo", value=estudiante.nombre)
                        nuevo_ci = st.text_input("Cédula de identidad", value=estudiante.ci)
                    
                    with col2:
                        nuevo_grupo = st.text_input("Grupo", value=estudiante.grupo or "")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        if st.form_submit_button("Actualizar Estudiante", type="primary", use_container_width=True):
                            if not nuevo_nombre or not nuevo_ci:
                                st.error("Nombre y CI son obligatorios")
                            else:
                                resultado = EstudianteManager.actualizar_estudiante(
                                    estudiante_id,
                                    nombre=nuevo_nombre,
                                    ci=nuevo_ci,
                                    grupo=nuevo_grupo or None
                                )
                                
                                if resultado:
                                    st.success("Estudiante actualizado exitosamente")
                                    del st.session_state['editar_estudiante_id']
                                    st.rerun()
                                else:
                                    st.error("No se pudo actualizar el estudiante")
                    
                    with col2:
                        if st.form_submit_button("Cancelar", use_container_width=True):
                            del st.session_state['editar_estudiante_id']
                            st.rerun()
    
    with tab3:
        st.subheader("Organizar Grupos Automáticamente")
        
        # Obtener paralelos
        paralelos_disponibles = []
        materias = MateriaManager.listar_materias()
        
        for materia in materias:
            paralelos = ParaleloManager.listar_paralelos_por_materia(materia.id)
            for paralelo in paralelos:
                paralelos_disponibles.append({
                    'texto': f"{materia.sigla} - Paralelo {paralelo.paralelo}",
                    'paralelo_id': paralelo.id
                })
        
        if paralelos_disponibles:
            with st.form("form_organizar_grupos"):
                opciones_paralelos = {p['texto']: p['paralelo_id'] for p in paralelos_disponibles}
                paralelo_seleccionado = st.selectbox(
                    "Paralelo:",
                    options=list(opciones_paralelos.keys())
                )
                
                estudiantes_por_grupo = st.number_input(
                    "Estudiantes por grupo:",
                    min_value=2,
                    max_value=10,
                    value=5
                )
                
                submitted = st.form_submit_button("Organizar Grupos", 
                                                type="primary",
                                                use_container_width=True)
                
                if submitted:
                    paralelo_id = opciones_paralelos[paralelo_seleccionado]
                    resultado = EstudianteManager.organizar_grupos_automatico(paralelo_id, estudiantes_por_grupo)
                    
                    if resultado['success']:
                        st.success(resultado['mensaje'])
                        st.info(f"Se crearon {resultado['grupos_creados']} grupos")
                    else:
                        st.error(resultado['mensaje'])
        else:
            st.warning("No hay paralelos registrados.")
    
    with tab4:
        st.subheader("Búsqueda de Estudiantes")

        ci_busqueda = st.text_input("Buscar por CI:", placeholder="Ingrese cédula de identidad")

        if ci_busqueda:
            estudiantes = EstudianteManager.buscar_todos_por_ci(ci_busqueda)

            if estudiantes:
                # Mostrar información general
                st.success(f"Se encontraron {len(estudiantes)} inscripción(es) para el CI: {ci_busqueda}")

                # Si es el mismo estudiante, mostrar nombre una sola vez
                if estudiantes:
                    st.write(f"**Nombre:** {estudiantes[0].nombre}")
                    st.write(f"**CI:** {estudiantes[0].ci}")

                st.divider()

                # Mostrar cada inscripción (paralelo/materia)
                for idx, estudiante in enumerate(estudiantes, 1):
                    st.subheader(f"Inscripción {idx}: {estudiante.id_paralelo.id_materia.sigla} - Paralelo {estudiante.id_paralelo.paralelo}")

                    col1, col2, col3 = st.columns(3)

                    with col1:
                        st.write(f"**Materia:** {estudiante.id_paralelo.id_materia.materia}")
                        st.write(f"**Paralelo:** {estudiante.id_paralelo.paralelo}")
                        st.write(f"**Docente:** {estudiante.id_paralelo.docente_teoria}")

                    with col2:
                        st.write(f"**Grupo:** {estudiante.grupo or 'Sin asignar'}")
                        st.write(f"**Promedio:** {estudiante.promedio_calificaciones():.2f}")
                        st.write(f"**Calificaciones:** {estudiante.contar_calificaciones()}")

                    with col3:
                        # Botones de acción para esta inscripción
                        if st.button(f"Ver detalles", key=f"ver_detalles_{estudiante.id}"):
                            st.session_state[f'mostrar_detalles_{estudiante.id}'] = not st.session_state.get(f'mostrar_detalles_{estudiante.id}', False)

                    # Mostrar calificaciones si se solicitó
                    if st.session_state.get(f'mostrar_detalles_{estudiante.id}', False):
                        calificaciones = estudiante.calificaciones_por_laboratorio()

                        if calificaciones:
                            st.write("**Calificaciones por Laboratorio:**")

                            datos_cal = []
                            for lab, datos in calificaciones.items():
                                datos_cal.append({
                                    'Laboratorio': lab,
                                    'Título': datos['titulo'],
                                    'Calificación': f"{datos['calificacion']:.1f}" if datos['calificacion'] is not None else 'Pendiente',
                                    'Fecha': datos['fecha'].strftime('%d/%m/%Y') if datos['fecha'] else 'N/A'
                                })

                            df_cal = pd.DataFrame(datos_cal)
                            st.dataframe(df_cal, use_container_width=True, hide_index=True)
                        else:
                            st.info("No tiene calificaciones registradas en esta materia")

                    st.divider()
            else:
                st.warning(f"No se encontró estudiante con CI: {ci_busqueda}")

def pagina_laboratorios():
    """Página de gestión de laboratorios"""
    st.header("Gestión de Laboratorios")
    
    tab1, tab2, tab3 = st.tabs(["Lista de Laboratorios", "Nuevo Laboratorio", "Estadísticas"])
    
    with tab1:
        st.subheader("Laboratorios por Materia")
        
        materias = MateriaManager.listar_materias()
        
        if not materias:
            st.warning("No hay materias registradas. Debe crear materias primero.")
            return
        
        opciones_materias = {f"{m.sigla} - {m.materia}": m.id for m in materias}
        materia_seleccionada = st.selectbox(
            "Seleccionar materia:",
            options=list(opciones_materias.keys())
        )
        
        if materia_seleccionada:
            materia_id = opciones_materias[materia_seleccionada]
            laboratorios = LaboratorioManager.listar_laboratorios_por_materia(materia_id)
            
            if laboratorios:
                datos = []
                for lab in laboratorios:
                    datos.append({
                        'ID': lab.id,
                        'Número': lab.numero,
                        'Título': lab.titulo,
                        'Puntaje Máximo': f"{lab.puntaje_maximo:.1f}",
                        'Calificaciones': lab.contar_calificaciones(),
                        'Promedio': f"{lab.promedio_calificaciones():.2f}"
                    })
                
                df = pd.DataFrame(datos)
                st.dataframe(df, use_container_width=True, hide_index=True)
                
                # Acciones sobre laboratorios
                st.subheader("Acciones sobre Laboratorios")
                
                opciones_labs = {f"Lab {l.numero}: {l.titulo}": l.id for l in laboratorios}
                lab_seleccionado = st.selectbox(
                    "Seleccionar laboratorio:",
                    options=list(opciones_labs.keys())
                )
                
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    if st.button("Ver Estadísticas", use_container_width=True):
                        lab_id = opciones_labs[lab_seleccionado]
                        lab = LaboratorioManager.obtener_laboratorio(lab_id)
                        
                        if lab:
                            stats = lab.estadisticas_detalladas()
                            
                            st.info(f"""
                            **Laboratorio:** {lab.titulo}
                            **Total calificaciones:** {stats['total_calificaciones']}
                            **Promedio:** {stats['promedio']:.2f}
                            **Nota máxima:** {stats['nota_maxima']:.1f}
                            **Nota mínima:** {stats['nota_minima']:.1f}
                            **Aprobados:** {stats['aprobados']}
                            **Reprobados:** {stats['reprobados']}
                            """)
                
                with col2:
                    if st.button("Editar Laboratorio", use_container_width=True):
                        lab_id = opciones_labs[lab_seleccionado]
                        st.session_state['editar_laboratorio_id'] = lab_id
                
                with col3:
                    # Manejo de estado de confirmación
                    if 'confirmar_eliminar_laboratorio' not in st.session_state:
                        st.session_state.confirmar_eliminar_laboratorio = None

                    lab_id = opciones_labs[lab_seleccionado]

                    if st.session_state.confirmar_eliminar_laboratorio == lab_id:
                        st.warning("⚠️ ¿Está seguro de eliminar este laboratorio?")
                        col_si, col_no = st.columns(2)

                        with col_si:
                            if st.button("✓ Sí, eliminar", type="primary", use_container_width=True, key=f"conf_lab_si_{lab_id}"):
                                resultado = LaboratorioManager.eliminar_laboratorio(lab_id, forzar=True)

                                if resultado['success']:
                                    st.success(resultado['mensaje'])
                                    st.session_state.confirmar_eliminar_laboratorio = None
                                    st.rerun()
                                else:
                                    st.error(resultado['mensaje'])
                                    st.session_state.confirmar_eliminar_laboratorio = None

                        with col_no:
                            if st.button("✗ Cancelar", use_container_width=True, key=f"conf_lab_no_{lab_id}"):
                                st.session_state.confirmar_eliminar_laboratorio = None
                                st.rerun()
                    else:
                        if st.button("Eliminar Laboratorio", type="primary", use_container_width=True):
                            st.session_state.confirmar_eliminar_laboratorio = lab_id
                            st.rerun()
            else:
                st.info("No hay laboratorios registrados para esta materia")
    
    with tab2:
        st.subheader("Crear Nuevo Laboratorio")
        
        materias = MateriaManager.listar_materias()
        
        if materias:
            with st.form("form_nuevo_laboratorio"):
                opciones_materias = {f"{m.sigla} - {m.materia}": m.id for m in materias}
                materia_seleccionada = st.selectbox(
                    "Materia:",
                    options=list(opciones_materias.keys())
                )
                
                titulo = st.text_input("Título del laboratorio *", placeholder="Ej: Variables y Tipos de Datos")
                descripcion = st.text_area("Descripción (opcional)", placeholder="Descripción detallada del laboratorio")
                puntaje_maximo = st.number_input("Puntaje máximo", min_value=1.0, value=100.0, step=0.5)
                
                submitted = st.form_submit_button("Crear Laboratorio", 
                                                type="primary",
                                                use_container_width=True)
                
                if submitted:
                    if not titulo:
                        st.error("El título es obligatorio")
                    else:
                        materia_id = opciones_materias[materia_seleccionada]
                        resultado = LaboratorioManager.crear_laboratorio(
                            materia_id, titulo, descripcion or None, puntaje_maximo
                        )
                        
                        if resultado:
                            st.success(f"Laboratorio '{titulo}' creado exitosamente")
                            st.rerun()
                        else:
                            st.error("No se pudo crear el laboratorio")
        else:
            st.warning("No hay materias registradas. Debe crear materias primero.")
        
        # Formulario de edición si hay un laboratorio seleccionado
        if 'editar_laboratorio_id' in st.session_state:
            st.subheader("Editar Laboratorio")
            
            lab_id = st.session_state['editar_laboratorio_id']
            lab = LaboratorioManager.obtener_laboratorio(lab_id)
            
            if lab:
                with st.form("form_editar_laboratorio"):
                    nuevo_titulo = st.text_input("Título del laboratorio", value=lab.titulo)
                    nueva_descripcion = st.text_area("Descripción", value=lab.descripcion or "")
                    nuevo_puntaje = st.number_input("Puntaje máximo", value=float(lab.puntaje_maximo), step=0.5)
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        if st.form_submit_button("Actualizar Laboratorio", type="primary", use_container_width=True):
                            if not nuevo_titulo:
                                st.error("El título es obligatorio")
                            else:
                                resultado = LaboratorioManager.actualizar_laboratorio(
                                    lab_id,
                                    titulo=nuevo_titulo,
                                    descripcion=nueva_descripcion or None,
                                    puntaje_maximo=nuevo_puntaje
                                )
                                
                                if resultado:
                                    st.success("Laboratorio actualizado exitosamente")
                                    del st.session_state['editar_laboratorio_id']
                                    st.rerun()
                                else:
                                    st.error("No se pudo actualizar el laboratorio")
                    
                    with col2:
                        if st.form_submit_button("Cancelar", use_container_width=True):
                            del st.session_state['editar_laboratorio_id']
                            st.rerun()
    
    with tab3:
        st.subheader("Estadísticas de Laboratorios")
        
        materias = MateriaManager.listar_materias()
        
        for materia in materias:
            laboratorios = LaboratorioManager.listar_laboratorios_por_materia(materia.id)
            
            if laboratorios:
                st.write(f"**{materia.sigla} - {materia.materia}**")
                
                datos_stats = []
                for lab in laboratorios:
                    stats = lab.estadisticas_detalladas()
                    datos_stats.append({
                        'Lab': lab.numero,
                        'Título': lab.titulo,
                        'Calificaciones': stats['total_calificaciones'],
                        'Promedio': f"{stats['promedio']:.2f}",
                        'Aprobados': stats['aprobados'],
                        'Reprobados': stats['reprobados']
                    })
                
                df_stats = pd.DataFrame(datos_stats)
                st.dataframe(df_stats, use_container_width=True, hide_index=True)
                
                st.markdown("---")

def pagina_calificaciones():
    """Página de gestión de calificaciones"""
    st.header("Gestión de Calificaciones")

    tab1, tab2, tab3, tab4, tab5 = st.tabs(["Lista de Calificaciones", "Nueva Calificación", "Calificar por Lotes", "Estadísticas", "Calificaciones por Materia"])
    
    with tab1:
        st.subheader("Calificaciones por Laboratorio")
        
        # Obtener laboratorios
        laboratorios_disponibles = []
        materias = MateriaManager.listar_materias()
        
        for materia in materias:
            laboratorios = LaboratorioManager.listar_laboratorios_por_materia(materia.id)
            for lab in laboratorios:
                laboratorios_disponibles.append({
                    'texto': f"{materia.sigla} - Lab {lab.numero}: {lab.titulo}",
                    'lab_id': lab.id,
                    'puntaje_max': lab.puntaje_maximo
                })
        
        if not laboratorios_disponibles:
            st.warning("No hay laboratorios registrados. Debe crear laboratorios primero.")
            return
        
        opciones_labs = {l['texto']: l for l in laboratorios_disponibles}
        lab_seleccionado = st.selectbox(
            "Seleccionar laboratorio:",
            options=list(opciones_labs.keys())
        )
        
        if lab_seleccionado:
            lab_info = opciones_labs[lab_seleccionado]
            lab_id = lab_info['lab_id']
            
            st.info(f"Puntaje máximo: {lab_info['puntaje_max']}")
            
            calificaciones = CalificacionManager.obtener_calificaciones_laboratorio(lab_id)
            
            if calificaciones:
                datos = []
                for cal in calificaciones:
                    estudiante = cal.id_estudiante
                    nota_str = f"{cal.calificacion:.1f}" if cal.calificacion else "Sin nota"
                    estado = cal.estado_aprobacion()
                    fecha = cal.fecha_registro.strftime("%d/%m/%Y") if cal.fecha_registro else "N/A"
                    
                    datos.append({
                        'ID': cal.id,
                        'CI': estudiante.ci,
                        'Estudiante': estudiante.nombre,
                        'Grupo': estudiante.grupo or "Sin asignar",
                        'Calificación': nota_str,
                        'Estado': estado,
                        'Fecha': fecha,
                        'Observaciones': cal.observaciones or ""
                    })
                
                df = pd.DataFrame(datos)
                st.dataframe(df, use_container_width=True, hide_index=True)
                
                # Acciones sobre calificaciones
                st.subheader("Acciones sobre Calificaciones")
                
                opciones_cal = {f"{d['CI']} - {d['Estudiante']}": d['ID'] for d in datos}
                cal_seleccionada = st.selectbox(
                    "Seleccionar calificación:",
                    options=list(opciones_cal.keys())
                )
                
                col1, col2 = st.columns(2)
                
                with col1:
                    if st.button("Editar Calificación", use_container_width=True):
                        cal_id = opciones_cal[cal_seleccionada]
                        st.session_state['editar_calificacion_id'] = cal_id
                        st.session_state['laboratorio_actual_id'] = lab_id
                
                with col2:
                    # Manejo de estado de confirmación
                    if 'confirmar_eliminar_calificacion' not in st.session_state:
                        st.session_state.confirmar_eliminar_calificacion = None

                    cal_id = opciones_cal[cal_seleccionada]

                    if st.session_state.confirmar_eliminar_calificacion == cal_id:
                        st.warning("⚠️ ¿Está seguro de eliminar esta calificación?")
                        col_si, col_no = st.columns(2)

                        with col_si:
                            if st.button("✓ Sí, eliminar", type="primary", use_container_width=True, key=f"conf_cal_si_{cal_id}"):
                                resultado = CalificacionManager.eliminar_calificacion(cal_id)

                                if resultado['success']:
                                    st.success(resultado['mensaje'])
                                    st.session_state.confirmar_eliminar_calificacion = None
                                    st.rerun()
                                else:
                                    st.error(resultado['mensaje'])
                                    st.session_state.confirmar_eliminar_calificacion = None

                        with col_no:
                            if st.button("✗ Cancelar", use_container_width=True, key=f"conf_cal_no_{cal_id}"):
                                st.session_state.confirmar_eliminar_calificacion = None
                                st.rerun()
                    else:
                        if st.button("Eliminar Calificación", type="primary", use_container_width=True):
                            st.session_state.confirmar_eliminar_calificacion = cal_id
                            st.rerun()
            else:
                st.info("No hay calificaciones registradas para este laboratorio")
    
    with tab2:
        st.subheader("Registrar Nueva Calificación")
        
        # Obtener laboratorios
        laboratorios_disponibles = []
        materias = MateriaManager.listar_materias()
        
        for materia in materias:
            laboratorios = LaboratorioManager.listar_laboratorios_por_materia(materia.id)
            for lab in laboratorios:
                laboratorios_disponibles.append({
                    'texto': f"{materia.sigla} - Lab {lab.numero}: {lab.titulo}",
                    'lab_id': lab.id,
                    'puntaje_max': lab.puntaje_maximo
                })
        
        if laboratorios_disponibles:
            with st.form("form_nueva_calificacion"):
                opciones_labs = {l['texto']: l for l in laboratorios_disponibles}
                lab_seleccionado = st.selectbox(
                    "Laboratorio:",
                    options=list(opciones_labs.keys())
                )
                
                lab_info = opciones_labs[lab_seleccionado]
                st.info(f"Puntaje máximo: {lab_info['puntaje_max']}")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    ci_estudiante = st.text_input("CI del estudiante *", placeholder="12345678")
                    calificacion = st.number_input("Calificación *", min_value=0.0, 
                                                 max_value=float(lab_info['puntaje_max']), 
                                                 step=0.1)
                
                with col2:
                    observaciones = st.text_area("Observaciones (opcional)", 
                                               placeholder="Comentarios sobre la calificación")
                
                submitted = st.form_submit_button("Registrar Calificación", 
                                                type="primary",
                                                use_container_width=True)
                
                if submitted:
                    if not ci_estudiante:
                        st.error("El CI del estudiante es obligatorio")
                    else:
                        estudiante = EstudianteManager.buscar_por_ci(ci_estudiante)
                        
                        if not estudiante:
                            st.error(f"No existe estudiante con CI: {ci_estudiante}")
                        else:
                            resultado = CalificacionManager.registrar_calificacion(
                                lab_info['lab_id'], estudiante.id, calificacion, observaciones or None
                            )
                            
                            if resultado:
                                st.success(f"Calificación registrada para {estudiante.nombre}")
                                st.rerun()
                            else:
                                st.error("No se pudo registrar la calificación")
        else:
            st.warning("No hay laboratorios registrados. Debe crear laboratorios primero.")
        
        # Formulario de edición si hay una calificación seleccionada
        if 'editar_calificacion_id' in st.session_state and 'laboratorio_actual_id' in st.session_state:
            st.subheader("Editar Calificación")
            
            cal_id = st.session_state['editar_calificacion_id']
            lab_id = st.session_state['laboratorio_actual_id']
            
            # Obtener calificación actual
            from models.calificacion import Calificacion
            try:
                cal = Calificacion.get_by_id(cal_id)
                lab = LaboratorioManager.obtener_laboratorio(lab_id)
                
                with st.form("form_editar_calificacion"):
                    st.info(f"Editando calificación de: {cal.id_estudiante.nombre} ({cal.id_estudiante.ci})")
                    st.info(f"Laboratorio: {lab.titulo} (Puntaje máximo: {lab.puntaje_maximo})")
                    
                    nueva_calificacion = st.number_input("Calificación", 
                                                       value=float(cal.calificacion) if cal.calificacion else 0.0,
                                                       min_value=0.0,
                                                       max_value=float(lab.puntaje_maximo),
                                                       step=0.1)
                    
                    nuevas_observaciones = st.text_area("Observaciones", 
                                                      value=cal.observaciones or "")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        if st.form_submit_button("Actualizar Calificación", type="primary", use_container_width=True):
                            resultado = CalificacionManager.actualizar_calificacion(
                                cal_id, nueva_calificacion, nuevas_observaciones or None
                            )
                            
                            if resultado:
                                st.success("Calificación actualizada exitosamente")
                                del st.session_state['editar_calificacion_id']
                                del st.session_state['laboratorio_actual_id']
                                st.rerun()
                            else:
                                st.error("No se pudo actualizar la calificación")
                    
                    with col2:
                        if st.form_submit_button("Cancelar", use_container_width=True):
                            del st.session_state['editar_calificacion_id']
                            del st.session_state['laboratorio_actual_id']
                            st.rerun()
                            
            except Exception as e:
                st.error(f"Error al cargar calificación: {e}")
                del st.session_state['editar_calificacion_id']
                del st.session_state['laboratorio_actual_id']
    
    with tab3:
        st.subheader("Calificar por Lotes")
        
        # Obtener laboratorios
        laboratorios_disponibles = []
        materias = MateriaManager.listar_materias()
        
        for materia in materias:
            laboratorios = LaboratorioManager.listar_laboratorios_por_materia(materia.id)
            for lab in laboratorios:
                laboratorios_disponibles.append({
                    'texto': f"{materia.sigla} - Lab {lab.numero}: {lab.titulo}",
                    'lab_id': lab.id,
                    'puntaje_max': lab.puntaje_maximo
                })
        
        if laboratorios_disponibles:
            with st.form("form_calificar_lotes"):
                opciones_labs = {l['texto']: l for l in laboratorios_disponibles}
                lab_seleccionado = st.selectbox(
                    "Laboratorio:",
                    options=list(opciones_labs.keys())
                )
                
                lab_info = opciones_labs[lab_seleccionado]
                st.info(f"Puntaje máximo: {lab_info['puntaje_max']}")
                
                st.write("Ingrese las calificaciones en el siguiente formato:")
                st.code("CI,calificacion\n12345678,85.5\n87654321,92.0", language="text")
                
                calificaciones_texto = st.text_area(
                    "Calificaciones por lotes:",
                    height=200,
                    placeholder="12345678,85.5\n87654321,92.0\n11223344,78.0"
                )
                
                submitted = st.form_submit_button("Procesar Calificaciones", 
                                                type="primary",
                                                use_container_width=True)
                
                if submitted:
                    if not calificaciones_texto.strip():
                        st.error("Debe ingresar al menos una calificación")
                    else:
                        # Procesar texto
                        lineas = [linea.strip() for linea in calificaciones_texto.strip().split('\n') if linea.strip()]
                        calificaciones_dict = {}
                        errores = []
                        
                        for i, linea in enumerate(lineas, 1):
                            if ',' not in linea:
                                errores.append(f"Línea {i}: Formato incorrecto")
                                continue
                            
                            try:
                                ci, calificacion_str = linea.split(',', 1)
                                ci = ci.strip()
                                calificacion = float(calificacion_str.strip())
                                
                                # Verificar estudiante
                                estudiante = EstudianteManager.buscar_por_ci(ci)
                                if not estudiante:
                                    errores.append(f"Línea {i}: No existe estudiante con CI {ci}")
                                    continue
                                
                                calificaciones_dict[estudiante.id] = calificacion
                                
                            except ValueError:
                                errores.append(f"Línea {i}: La calificación debe ser un número")
                                continue
                            except Exception as e:
                                errores.append(f"Línea {i}: {e}")
                                continue
                        
                        if calificaciones_dict:
                            resultado = CalificacionManager.calificar_por_lotes(lab_info['lab_id'], calificaciones_dict)
                            
                            if resultado['success']:
                                st.success(f"Procesamiento completado: {resultado['exitosas']} calificaciones registradas")
                                
                                if resultado['errores']:
                                    st.warning(f"Se encontraron {len(resultado['errores'])} errores:")
                                    for error in resultado['errores'][:5]:  # Mostrar solo los primeros 5
                                        st.write(f"- {error}")
                                    if len(resultado['errores']) > 5:
                                        st.write(f"... y {len(resultado['errores']) - 5} errores más")
                            else:
                                st.error(resultado['mensaje'])
                        
                        if errores:
                            st.error(f"Errores en el formato:")
                            for error in errores[:10]:  # Mostrar solo los primeros 10
                                st.write(f"- {error}")
                            if len(errores) > 10:
                                st.write(f"... y {len(errores) - 10} errores más")
                        
                        if not calificaciones_dict and not errores:
                            st.warning("No se procesaron calificaciones válidas")
        else:
            st.warning("No hay laboratorios registrados. Debe crear laboratorios primero.")
    
    with tab4:
        st.subheader("Estadísticas de Calificaciones")
        
        # Obtener paralelos para estadísticas
        paralelos_disponibles = []
        materias = MateriaManager.listar_materias()
        
        for materia in materias:
            paralelos = ParaleloManager.listar_paralelos_por_materia(materia.id)
            for paralelo in paralelos:
                paralelos_disponibles.append({
                    'texto': f"{materia.sigla} - Paralelo {paralelo.paralelo}",
                    'paralelo_id': paralelo.id
                })
        
        if paralelos_disponibles:
            opciones_paralelos = {p['texto']: p['paralelo_id'] for p in paralelos_disponibles}
            paralelo_seleccionado = st.selectbox(
                "Seleccionar paralelo para estadísticas:",
                options=list(opciones_paralelos.keys())
            )
            
            if paralelo_seleccionado:
                paralelo_id = opciones_paralelos[paralelo_seleccionado]
                paralelo = ParaleloManager.obtener_paralelo(paralelo_id)
                
                from models.calificacion import Calificacion
                stats = Calificacion.estadisticas_paralelo(paralelo)
                
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    st.metric("Total Calificaciones", stats['total_calificaciones'])
                
                with col2:
                    st.metric("Promedio General", f"{stats['promedio_general']:.2f}")
                
                with col3:
                    st.metric("Aprobados", stats['aprobados'])
                
                with col4:
                    st.metric("Reprobados", stats['reprobados'])
                
                if stats['total_calificaciones'] > 0:
                    porcentaje_aprobacion = (stats['aprobados'] / stats['total_calificaciones']) * 100
                    st.metric("Porcentaje de Aprobación", f"{porcentaje_aprobacion:.1f}%")
                
                if stats['sin_calificar'] > 0:
                    st.warning(f"Hay {stats['sin_calificar']} registros sin calificar")
        else:
            st.info("No hay paralelos registrados para mostrar estadísticas.")

    with tab5:
        st.subheader("Calificaciones por Materia")

        # Obtener materias
        materias = MateriaManager.listar_materias()

        if not materias:
            st.warning("No hay materias registradas.")
            return

        opciones_materias = {f"{materia.sigla} - {materia.materia}": materia.id for materia in materias}
        materia_seleccionada = st.selectbox(
            "Seleccionar materia:",
            options=list(opciones_materias.keys()),
            key="materia_calificaciones"
        )

        if materia_seleccionada:
            materia_id = opciones_materias[materia_seleccionada]
            materia = MateriaManager.obtener_materia(materia_id)

            # Obtener todos los paralelos de la materia
            paralelos = ParaleloManager.listar_paralelos_por_materia(materia_id)

            if not paralelos:
                st.info("No hay paralelos registrados para esta materia.")
                return

            # Obtener todos los laboratorios de la materia
            laboratorios = LaboratorioManager.listar_laboratorios_por_materia(materia_id)

            if not laboratorios:
                st.info("No hay laboratorios registrados para esta materia.")
                return

            # Mostrar info de la materia
            st.info(f"**{materia.sigla}** - {materia.materia}")
            st.write(f"**Paralelos:** {len(paralelos)} | **Laboratorios:** {len(laboratorios)}")

            # Crear matriz de calificaciones para todos los estudiantes en la materia
            from models.calificacion import Calificacion
            from models.laboratorio import Laboratorio

            # Obtener estudiantes de todos los paralelos de la materia
            estudiantes_materia = []
            for paralelo in paralelos:
                estudiantes_paralelo = EstudianteManager.listar_por_paralelo(paralelo.id)
                for est in estudiantes_paralelo:
                    estudiantes_materia.append((est, paralelo.paralelo))  # Incluir paralelo

            if not estudiantes_materia:
                st.info("No hay estudiantes registrados en esta materia.")
                return

            # Crear DataFrame con calificaciones por estudiante y laboratorio
            datos_matriz = []
            for estudiante, paralelo_nombre in estudiantes_materia:
                fila = {
                    'CI': estudiante.ci,
                    'Estudiante': estudiante.nombre,
                    'Paralelo': paralelo_nombre,
                    'Grupo': estudiante.grupo or 'Sin asignar'
                }

                # Agregar calificaciones para cada laboratorio
                for lab in laboratorios:
                    # Buscar calificación específica para este estudiante y laboratorio
                    calificacion = CalificacionManager.obtener_calificacion_especifica(lab.id, estudiante.id)
                    if calificacion and calificacion.calificacion is not None:
                        fila[f'Lab {lab.numero}'] = f"{calificacion.calificacion:.1f}"
                    else:
                        fila[f'Lab {lab.numero}'] = "--"

                # Calcular promedio basado en la cantidad total de laboratorios
                calificaciones = []
                for lab in laboratorios:
                    calificacion = CalificacionManager.obtener_calificacion_especifica(lab.id, estudiante.id)
                    if calificacion and calificacion.calificacion is not None:
                        calificaciones.append(calificacion.calificacion)

                suma_calificaciones = sum(calificaciones) if calificaciones else 0
                promedio = round(suma_calificaciones / len(laboratorios), 2) if len(laboratorios) > 0 else 0.0
                fila['Promedio'] = f"{promedio:.2f}"

                datos_matriz.append(fila)

            df_matriz = pd.DataFrame(datos_matriz)

            # Mostrar formulario vertical para editar calificaciones de un estudiante específico
            st.subheader("Editar Calificaciones por Estudiante")

            # Crear diccionario de estudiantes para el selector
            estudiantes_dict = {}
            for est, par in estudiantes_materia:
                estudiantes_dict[f"{est.ci} - {est.nombre} (Paralelo {par})"] = (est, par)

            if estudiantes_dict:
                # Selector de estudiante
                estudiante_seleccionado = st.selectbox(
                    "Seleccionar estudiante:",
                    options=list(estudiantes_dict.keys()),
                    key="select_estudiante_materia"
                )

                if estudiante_seleccionado:
                    estudiante, paralelo_nombre = estudiantes_dict[estudiante_seleccionado]

                    # Mostrar información del estudiante
                    st.info(f"**Estudiante:** {estudiante.nombre}  \n**CI:** {estudiante.ci}  \n**Paralelo:** {paralelo_nombre}")

                    # Crear formulario vertical para editar todas las calificaciones
                    with st.form(f"form_calificaciones_{estudiante.id}"):
                        # Diccionario para almacenar los valores actuales
                        valores_calificaciones = {}

                        for lab in laboratorios:
                            # Obtener la calificación actual
                            calificacion = CalificacionManager.obtener_calificacion_especifica(lab.id, estudiante.id)
                            calificacion_valor = calificacion.calificacion if calificacion and calificacion.calificacion is not None else None

                            # Campo para editar la calificación
                            nuevo_valor = st.number_input(
                                f"Lab {lab.numero}: {lab.titulo}",
                                value=calificacion_valor if calificacion_valor is not None else 0.0,
                                min_value=0.0,
                                max_value=lab.puntaje_maximo,
                                step=0.1,
                                key=f"cal_input_{estudiante.id}_{lab.id}",
                                help=f"Puntaje máximo: {lab.puntaje_maximo}"
                            )

                            valores_calificaciones[lab.id] = {
                                'actual': calificacion_valor,
                                'nuevo': nuevo_valor,
                                'objeto': calificacion
                            }

                        # Botón para guardar todas las calificaciones
                        submitted = st.form_submit_button("Guardar todas las calificaciones", type="primary")

                        if submitted:
                            actualizaciones_realizadas = 0
                            for lab_id, data in valores_calificaciones.items():
                                nuevo_valor = data['nuevo']
                                actual_valor = data['actual']
                                calificacion = data['objeto']

                                # Solo actualizar si hay un cambio real
                                hay_cambio = False
                                if actual_valor is None and nuevo_valor != 0.0:
                                    hay_cambio = True
                                elif actual_valor is not None and abs(nuevo_valor - actual_valor) > 0.01:
                                    hay_cambio = True

                                if hay_cambio:
                                    lab = next(l for l in laboratorios if l.id == lab_id)
                                    if calificacion:
                                        # Actualizar existente
                                        CalificacionManager.actualizar_calificacion(calificacion.id, nuevo_valor)
                                    else:
                                        # Crear nueva calificación
                                        CalificacionManager.registrar_calificacion(lab.id, estudiante.id, nuevo_valor)
                                    actualizaciones_realizadas += 1

                            if actualizaciones_realizadas > 0:
                                st.success(f"✓ {actualizaciones_realizadas} calificación(es) actualizada(s) para {estudiante.nombre}")
                                st.rerun()
                            else:
                                st.info("No hubo cambios para guardar")
            else:
                st.info("No hay estudiantes registrados en esta materia.")

            # Mostrar tabla original (puede ser útil para referencia)
            st.subheader("Vista de Calificaciones")
            st.dataframe(df_matriz, use_container_width=True, hide_index=True)

            # Opción para exportar a Excel
            if st.button("Exportar a Excel", use_container_width=True):
                try:
                    import io
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment
                    from openpyxl.utils.dataframe import dataframe_to_rows

                    # Crear archivo Excel en memoria
                    output = io.BytesIO()
                    wb = Workbook()
                    ws = wb.active
                    ws.title = f"{materia.sigla}_calificaciones"

                    # Agregar datos desde el DataFrame
                    for r in dataframe_to_rows(df_matriz, index=False, header=True):
                        ws.append(r)

                    # Formatear encabezado
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    center_alignment = Alignment(horizontal="center", vertical="center")

                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment

                    # Autoajustar columnas
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Limitar ancho máximo
                        ws.column_dimensions[column_letter].width = adjusted_width

                    wb.save(output)
                    output.seek(0)

                    # Guardar en el sistema de archivos
                    import os
                    from datetime import datetime
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"calificaciones_{materia.sigla}_{timestamp}.xlsx"
                    filepath = os.path.join("exports", "excel", filename)

                    # Crear directorio si no existe
                    os.makedirs(os.path.dirname(filepath), exist_ok=True)

                    # Guardar archivo
                    with open(filepath, "wb") as f:
                        f.write(output.getvalue())

                    st.success(f"Archivo Excel exportado: {filename}")
                    st.download_button(
                        label="Descargar Excel",
                        data=output.getvalue(),
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                except ImportError:
                    st.error("Para exportar a Excel, instale openpyxl: pip install openpyxl")
                except Exception as e:
                    st.error(f"Error al exportar a Excel: {e}")

def pagina_reportes():
    """Página de reportes y exportación"""
    st.header("Reportes y Exportación")
    
    tab1, tab2, tab3 = st.tabs(["Generar Reportes", "Matriz de Calificaciones", "Archivos Generados"])
    
    with tab1:
        st.subheader("Generar Reportes PDF")

        # Reporte consolidado (todas las materias y paralelos)
        st.markdown("### 📚 Reporte Consolidado")
        st.info("Genera un reporte simple de todas las materias y paralelos en un solo documento PDF")

        if st.button("📋 Generar Reporte Consolidado", type="primary", use_container_width=True, help="Reporte de todas las materias y paralelos"):
            with st.spinner("Generando reporte consolidado PDF..."):
                try:
                    archivo = PDFExporter.generar_reporte_consolidado()

                    if archivo:
                        st.success(f"Reporte consolidado PDF generado exitosamente!")
                        st.info(f"Ubicación: {archivo}")

                        # Mostrar botón de descarga si el archivo existe
                        if os.path.exists(archivo):
                            with open(archivo, "rb") as file:
                                st.download_button(
                                    label="📥 Descargar Reporte Consolidado",
                                    data=file.read(),
                                    file_name=os.path.basename(archivo),
                                    mime="application/pdf",
                                    use_container_width=True
                                )
                    else:
                        st.error("Error al generar el reporte consolidado")
                except Exception as e:
                    st.error(f"Error: {e}")

        st.markdown("---")
        st.markdown("### 📑 Reportes por Paralelo")

        # Obtener paralelos
        paralelos_disponibles = []
        materias = MateriaManager.listar_materias()
        
        for materia in materias:
            paralelos = ParaleloManager.listar_paralelos_por_materia(materia.id)
            for paralelo in paralelos:
                paralelos_disponibles.append({
                    'texto': f"{materia.sigla} - Paralelo {paralelo.paralelo}",
                    'paralelo_id': paralelo.id
                })
        
        if paralelos_disponibles:
            opciones_paralelos = {p['texto']: p['paralelo_id'] for p in paralelos_disponibles}
            paralelo_seleccionado = st.selectbox(
                "Seleccionar paralelo para reporte:",
                options=list(opciones_paralelos.keys())
            )
            
            col1, col2, col3 = st.columns(3)

            with col1:
                if st.button("📄 Reporte Simple", type="primary", use_container_width=True, help="Reporte para presentación con solo nombres y promedios"):
                    paralelo_id = opciones_paralelos[paralelo_seleccionado]

                    with st.spinner("Generando reporte simple PDF..."):
                        try:
                            archivo = PDFExporter.generar_reporte_simple(paralelo_id)

                            if archivo:
                                st.success(f"Reporte simple PDF generado exitosamente!")
                                st.info(f"Ubicación: {archivo}")

                                # Mostrar botón de descarga si el archivo existe
                                if os.path.exists(archivo):
                                    with open(archivo, "rb") as file:
                                        st.download_button(
                                            label="📥 Descargar Reporte Simple",
                                            data=file.read(),
                                            file_name=os.path.basename(archivo),
                                            mime="application/pdf",
                                            use_container_width=True
                                        )
                            else:
                                st.error("Error al generar el reporte PDF simple")
                        except Exception as e:
                            st.error(f"Error: {e}")

            with col2:
                if st.button("📊 Reporte Completo", use_container_width=True, help="Reporte detallado con matriz de calificaciones y estadísticas"):
                    paralelo_id = opciones_paralelos[paralelo_seleccionado]

                    with st.spinner("Generando reporte PDF completo..."):
                        try:
                            archivo = PDFExporter.generar_reporte_paralelo(paralelo_id)

                            if archivo:
                                st.success(f"Reporte completo PDF generado exitosamente!")
                                st.info(f"Ubicación: {archivo}")

                                # Mostrar botón de descarga si el archivo existe
                                if os.path.exists(archivo):
                                    with open(archivo, "rb") as file:
                                        st.download_button(
                                            label="📥 Descargar Reporte Completo",
                                            data=file.read(),
                                            file_name=os.path.basename(archivo),
                                            mime="application/pdf",
                                            use_container_width=True
                                        )
                            else:
                                st.error("Error al generar el reporte PDF completo")
                        except Exception as e:
                            st.error(f"Error: {e}")

            with col3:
                if st.button("📗 Reporte Excel", use_container_width=True):
                    st.info("Funcionalidad Excel próximamente disponible")
        else:
            st.warning("No hay paralelos registrados para generar reportes.")
    
    with tab2:
        st.subheader("Matriz de Calificaciones")
        
        # Obtener paralelos
        paralelos_disponibles = []
        materias = MateriaManager.listar_materias()
        
        for materia in materias:
            paralelos = ParaleloManager.listar_paralelos_por_materia(materia.id)
            for paralelo in paralelos:
                paralelos_disponibles.append({
                    'texto': f"{materia.sigla} - Paralelo {paralelo.paralelo}",
                    'paralelo_id': paralelo.id
                })
        
        if paralelos_disponibles:
            opciones_paralelos = {p['texto']: p['paralelo_id'] for p in paralelos_disponibles}
            paralelo_seleccionado = st.selectbox(
                "Seleccionar paralelo para matriz:",
                options=list(opciones_paralelos.keys()),
                key="matriz_paralelo"
            )
            
            if st.button("Mostrar Matriz", type="primary"):
                paralelo_id = opciones_paralelos[paralelo_seleccionado]
                
                try:
                    paralelo = ParaleloManager.obtener_paralelo(paralelo_id)
                    
                    from models.calificacion import Calificacion
                    from models.laboratorio import Laboratorio
                    
                    matriz = Calificacion.matriz_calificaciones_paralelo(paralelo)
                    laboratorios = list(Laboratorio.obtener_por_materia(paralelo.id_materia))
                    
                    if matriz and laboratorios:
                        st.write(f"**Matriz de Calificaciones: {paralelo.id_materia.sigla} - Paralelo {paralelo.paralelo}**")
                        
                        # Crear DataFrame para la matriz
                        datos_matriz = []
                        for fila in matriz:
                            row_data = {
                                'CI': fila['ci'],
                                'Estudiante': fila['estudiante'],
                                'Grupo': fila['grupo'] or 'Sin asignar'
                            }
                            
                            # Agregar calificaciones por laboratorio
                            for lab in laboratorios:
                                cal = fila['calificaciones'].get(f'lab_{lab.numero}')
                                row_data[f'Lab {lab.numero}'] = f"{cal:.1f}" if cal is not None else "--"
                            
                            row_data['Promedio'] = f"{fila['promedio']:.1f}"
                            datos_matriz.append(row_data)
                        
                        df_matriz = pd.DataFrame(datos_matriz)
                        st.dataframe(df_matriz, use_container_width=True, hide_index=True)
                        
                        # Estadísticas de la matriz
                        st.subheader("Estadísticas de la Matriz")
                        
                        from models.calificacion import Calificacion
                        stats = Calificacion.estadisticas_paralelo(paralelo)
                        
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            st.metric("Total Estudiantes", len(matriz))
                        
                        with col2:
                            st.metric("Promedio General", f"{stats['promedio_general']:.2f}")
                        
                        with col3:
                            porcentaje_aprobacion = (stats['aprobados'] / stats['total_calificaciones'] * 100) if stats['total_calificaciones'] > 0 else 0
                            st.metric("% Aprobación", f"{porcentaje_aprobacion:.1f}%")
                    else:
                        st.info("No hay datos suficientes para mostrar la matriz")
                        
                except Exception as e:
                    st.error(f"Error al generar matriz: {e}")
        else:
            st.warning("No hay paralelos registrados.")
    
    with tab3:
        st.subheader("Archivos Generados")
        
        carpeta_exports = "exports"
        
        if os.path.exists(carpeta_exports):
            archivos_encontrados = []
            
            # Buscar archivos PDF y Excel
            for root, dirs, files in os.walk(carpeta_exports):
                for file in files:
                    if file.endswith(('.pdf', '.xlsx', '.xls')):
                        archivo_path = os.path.join(root, file)
                        archivo_stat = os.stat(archivo_path)
                        
                        archivos_encontrados.append({
                            'Archivo': file,
                            'Tipo': 'PDF' if file.endswith('.pdf') else 'Excel',
                            'Tamaño (KB)': f"{archivo_stat.st_size / 1024:.1f}",
                            'Fecha': datetime.fromtimestamp(archivo_stat.st_mtime).strftime('%d/%m/%Y %H:%M'),
                            'Ruta': archivo_path
                        })
            
            if archivos_encontrados:
                df_archivos = pd.DataFrame(archivos_encontrados)
                st.dataframe(df_archivos[['Archivo', 'Tipo', 'Tamaño (KB)', 'Fecha']], 
                           use_container_width=True, hide_index=True)
                
                # Opciones de descarga
                st.subheader("Descargar Archivos")
                
                archivo_seleccionado = st.selectbox(
                    "Seleccionar archivo para descargar:",
                    options=[f"{a['Archivo']} ({a['Tipo']})" for a in archivos_encontrados]
                )
                
                if st.button("Descargar Archivo Seleccionado"):
                    # Encontrar el archivo seleccionado
                    nombre_archivo = archivo_seleccionado.split(' (')[0]
                    archivo_encontrado = next(a for a in archivos_encontrados if a['Archivo'] == nombre_archivo)
                    
                    try:
                        with open(archivo_encontrado['Ruta'], "rb") as file:
                            mime_type = "application/pdf" if archivo_encontrado['Tipo'] == 'PDF' else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            
                            st.download_button(
                                label=f"Descargar {nombre_archivo}",
                                data=file.read(),
                                file_name=nombre_archivo,
                                mime=mime_type
                            )
                    except Exception as e:
                        st.error(f"Error al descargar archivo: {e}")
            else:
                st.info("No hay archivos generados aún")
        else:
            st.info("No se han generado archivos aún")

def pagina_estadisticas():
    """Página de estadísticas generales"""
    st.header("Estadísticas Generales del Sistema")
    
    try:
        stats = MateriaManager.obtener_estadisticas_generales()
        materias = MateriaManager.listar_materias()
        
        # Resumen general
        st.subheader("Resumen General")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total Materias", stats['total_materias'])
        
        with col2:
            st.metric("Total Paralelos", stats['total_paralelos'])
        
        with col3:
            st.metric("Total Estudiantes", stats['total_estudiantes'])
        
        with col4:
            st.metric("Total Laboratorios", stats['total_laboratorios'])
        
        st.markdown("---")
        
        # Promedios
        st.subheader("Indicadores Promedio")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.metric("Paralelos por Materia", f"{stats['promedio_paralelos_por_materia']:.1f}")
        
        with col2:
            st.metric("Estudiantes por Materia", f"{stats['promedio_estudiantes_por_materia']:.1f}")
        
        st.markdown("---")
        
        # Detalle por materia
        st.subheader("Detalle por Materia")
        
        if materias:
            datos_detalle = []
            for materia in materias:
                stats_materia = materia.estadisticas_completas()
                datos_detalle.append({
                    'Sigla': stats_materia['sigla'],
                    'Materia': stats_materia['materia'],
                    'Paralelos': stats_materia['paralelos'],
                    'Estudiantes': stats_materia['estudiantes'],
                    'Laboratorios': stats_materia['laboratorios']
                })
            
            df_detalle = pd.DataFrame(datos_detalle)
            st.dataframe(df_detalle, use_container_width=True, hide_index=True)
            
            # Gráficos de estadísticas (usando los datos pero sin librerías de gráficos adicionales)
            st.subheader("Distribución de Datos")
            
            # Mostrar información en formato tabular ya que no tenemos matplotlib u otras librerías
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**Distribución de Estudiantes por Materia**")
                df_estudiantes = df_detalle[['Sigla', 'Estudiantes']].sort_values('Estudiantes', ascending=False)
                st.bar_chart(df_estudiantes.set_index('Sigla')['Estudiantes'])
            
            with col2:
                st.write("**Distribución de Laboratorios por Materia**")
                df_laboratorios = df_detalle[['Sigla', 'Laboratorios']].sort_values('Laboratorios', ascending=False)
                st.bar_chart(df_laboratorios.set_index('Sigla')['Laboratorios'])
        
        st.markdown("---")
        
        # Información adicional
        st.subheader("Información del Sistema")
        
        info_sistema = f"""
        **Sistema de Gestión de Laboratorios**
        
        **Estadísticas generadas:** {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
        
        **Resumen ejecutivo:**
        - Se han registrado {stats['total_materias']} materias en el sistema
        - Hay un total de {stats['total_paralelos']} paralelos distribuidos entre las materias
        - Se cuenta con {stats['total_estudiantes']} estudiantes inscritos
        - Se han definido {stats['total_laboratorios']} laboratorios para evaluación
        
        **Indicadores de eficiencia:**
        - Promedio de paralelos por materia: {stats['promedio_paralelos_por_materia']:.1f}
        - Promedio de estudiantes por materia: {stats['promedio_estudiantes_por_materia']:.1f}
        
        **Estado del sistema:** Operativo y funcional
        """
        
        st.markdown(info_sistema)
        
    except Exception as e:
        st.error(f"Error al cargar estadísticas: {e}")

def main():
    """Función principal de la aplicación Streamlit"""
    
    # Inicializar aplicación
    inicializar_aplicacion()
    
    # Mostrar título principal
    mostrar_titulo_principal()
    
    # Navegación principal
    pagina = sidebar_navegacion()
    
    # Renderizar página según selección
    if pagina == "Dashboard":
        mostrar_dashboard()
    elif pagina == "Materias":
        pagina_materias()
    elif pagina == "Paralelos":
        pagina_paralelos()
    elif pagina == "Estudiantes":
        pagina_estudiantes()
    elif pagina == "Laboratorios":
        pagina_laboratorios()
    elif pagina == "Calificaciones":
        pagina_calificaciones()
    elif pagina == "Reportes":
        pagina_reportes()
    elif pagina == "Estadísticas":
        pagina_estadisticas()
    
    # Footer
    st.markdown("---")
    st.markdown("""
    <div style='text-align: center; color: #666; font-size: 0.9em; padding: 20px;'>
        Sistema de Gestión de Laboratorios v1.0 - Universidad Técnica de Oruro<br>
        Desarrollado para la materia SIS 2420 - Actualización Tecnológica<br>
        Interfaz web desarrollada con Python + Streamlit
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
